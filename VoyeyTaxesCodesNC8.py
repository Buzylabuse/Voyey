


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import re
import xlsxwriter
from openpyxl import load_workbook
import xlrd
import numpy as np
import pandas as pd

book = xlrd.open_workbook('/Users/victorbuzy/Desktop/projetpython/Voyey/InputsVoyey.xls')
print(book)

sheet = book.sheet_by_index(0)

rows = sheet.nrows
cols = sheet.ncols
rows = int(rows)
print(rows)
print(cols)

myFileName = '/Users/victorbuzy/Desktop/projetpython/Voyey/VoyeyNC8.xlsx'
# load the workbook, and put the sheet into a variable
wb = load_workbook(filename=myFileName)
ws = wb['NC8']

Path = '/Users/victorbuzy/Desktop/projetpython/chromedriver'
options = webdriver.ChromeOptions()
# options.add_argument('headless')
options.add_argument('window-size=1200x600')  # optional

driver = webdriver.Chrome(Path)

for i in range(1, rows):
    # Start=(input("Quel article recherchez-vous?"))
    Start = sheet.cell_value(rowx=i, colx=0)
    # tout mettre en minuscule
    if Start.lower() in ["iphone12", "iphonex", "iphone12", "iphone10", "iphone11", "galaxys10", "galaxys9",
                         "galaxys8"]:
        Start = 'Telephone'
    elif Start.lower() in ['ps5', "ps4", "playstation4", "playstation5", "xbox", "xbox720", "xbox360"]:
        Start = 'Console de jeux'
    elif Start.lower() in ["macbook air", "macbook pro", "mac"]:
        Start = 'Ordinateur'
    elif Start.lower() in ["manettes", "manette", "manette de jeux"]:
        Start = 'Telecommande'

    print (Start)


    driver.get('https://google.com')
    time.sleep(2)

    search = driver.find_element_by_css_selector(
        'body > div.L3eUgb > div.o3j99.ikrT4e.om7nvf > form > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')

    search.send_keys('nc8 code SH ' + str(Start))
    search.send_keys(Keys.RETURN)

    time.sleep(5)
    NC8 = driver.find_element_by_partial_link_text('https://www.tarifdouanier.eu').text
    print(NC8)

    NC8 = re.findall('\d+', NC8)
    print(NC8)
    taille = len(NC8)
    if taille > 1:
        del NC8[1]
    NC8 = int("".join(str(i) for i in NC8))
    print(NC8)



    # max_row is a sheet function that gets the last row in a sheet.
    newRowLocation = ws.max_row + 1

    # write to the cell you want, specifying row and column, and value :-)
    ws.cell(column=1, row=newRowLocation, value=Start)
    ws.cell(column=2, row=newRowLocation, value=NC8)

wb.save(filename=myFileName)
wb.close()

driver.close()
driver.quit()
